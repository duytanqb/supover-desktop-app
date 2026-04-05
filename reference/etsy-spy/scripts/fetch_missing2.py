#!/usr/bin/env python3
import openpyxl, datetime, subprocess, json, sys, os, time

BACKLOG_PATH = "/Users/duytan/Documents/Business/Printfamily/etsy-ideas.xlsx"
SHEET = "backlog"

def load_backlog():
    wb = openpyxl.load_workbook(BACKLOG_PATH)
    ws = wb[SHEET]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        rows.append(dict(zip(headers, row)))
    return rows

def get_analytics(listing_id):
    cmd = [sys.executable, "etsy_analytics.py", "listing", str(listing_id)]
    try:
        out = subprocess.check_output(cmd, text=True, timeout=30, cwd=os.path.dirname(__file__))
        data = {}
        for line in out.split('\n'):
            if ':' in line:
                k, v = line.split(':', 1)
                data[k.strip()] = v.strip()
        # Map keys to backlog fields
        mapping = {
            'Recent Sold': 'Sold (24h)',
            'Views 24h': 'Views (24h)',
            'HEY Score': 'Hey',
            'Age (days)': 'Age (days)',
            'Total Sold': 'Total Sold',
            'Est. Revenue': 'Estimated Revenue',
            'Tags': 'Tags'
        }
        mapped = {}
        for k, v in data.items():
            for src, dst in mapping.items():
                if src in k:
                    mapped[dst] = v
                    break
        return mapped
    except Exception as e:
        print(f"  Analytics error: {e}")
        return {}

def update_row(row_id, updates):
    cmd = [sys.executable, "etsy_backlog.py", "update", "--ids", str(row_id)]
    for field, val in updates.items():
        if val is not None and val != "":
            cmd.extend([f"--{field.replace('_', '-')}", str(val)])
    try:
        subprocess.run(cmd, timeout=30, cwd=os.path.dirname(__file__), capture_output=True)
        return True
    except Exception as e:
        print(f"  Update error: {e}")
        return False

def main():
    rows = load_backlog()
    today = datetime.datetime.now().strftime("%Y-%m-%d")
    today_rows = [r for r in rows if r.get('spy_date') == today]
    print(f"Total rows today: {len(today_rows)}")
    
    updated = 0
    for row in today_rows:
        rid = row.get('id')
        lid = row.get('listing_id')
        sold = row.get('sold_24h')
        views = row.get('views_24h')
        tags = row.get('tags')
        # Check if missing
        if (sold in (None, '', 0, '0', '0.0') or 
            views in (None, '', 0, '0', '0.0') or 
            tags in (None, '')):
            print(f"\nRow {rid} (#{lid}) missing data")
            analytics = get_analytics(lid)
            if not analytics:
                print("  No analytics.")
                continue
            sold_new = analytics.get('Sold (24h)', '0')
            views_new = analytics.get('Views (24h)', '0')
            hey = analytics.get('Hey', '0')
            days = analytics.get('Age (days)', '0')
            total = analytics.get('Total Sold', '0')
            revenue = analytics.get('Estimated Revenue', '0')
            tags_new = analytics.get('Tags', '')
            # Convert
            try:
                sold_new_f = float(sold_new.replace(',', '')) if sold_new else 0.0
                views_new_f = float(views_new.replace(',', '')) if views_new else 0.0
                hey_f = float(hey) if hey else 0.0
                days_f = float(days) if days else 0.0
            except:
                sold_new_f = 0.0
                views_new_f = 0.0
                hey_f = 0.0
                days_f = 0.0
            # Determine if new data is better
            better = False
            updates = {}
            if sold_new_f > 0:
                updates['sold_24h'] = sold_new_f
                better = True
            if views_new_f > 0:
                updates['views_24h'] = views_new_f
                better = True
            if tags_new:
                updates['tags'] = tags_new[:200]
                better = True
            if hey_f > 0:
                updates['hey_score'] = hey_f
            if days_f > 0:
                updates['days_old'] = days_f
            if total != '0':
                updates['total_sold'] = total.replace(',', '')
            if revenue and revenue != '0':
                updates['estimated_revenue'] = revenue
            if better:
                print(f"  New data: sold_24h={sold_new_f}, views_24h={views_new_f}, tags={tags_new[:50]}...")
                if update_row(rid, updates):
                    updated += 1
                    print(f"  Updated row {rid}.")
                else:
                    print(f"  Failed to update row {rid}.")
            else:
                print(f"  No better data.")
            time.sleep(1)
    print(f"\n✅ Updated {updated} rows.")

if __name__ == "__main__":
    main()