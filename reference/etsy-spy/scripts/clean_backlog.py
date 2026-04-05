#!/usr/bin/env python3
"""
Clear all data in backlog sheet (keep header).
"""
import openpyxl
from openpyxl.utils import get_column_letter

BACKLOG_PATH = "/Users/duytan/Documents/Business/Printfamily/etsy-ideas.xlsx"
SHEET = "backlog"

def main():
    wb = openpyxl.load_workbook(BACKLOG_PATH)
    if SHEET not in wb.sheetnames:
        print(f"Sheet {SHEET} not found.")
        return
    ws = wb[SHEET]
    # Delete all rows except header (row 1)
    max_row = ws.max_row
    if max_row > 1:
        ws.delete_rows(2, max_row - 1)
        print(f"Deleted {max_row - 1} rows from {SHEET}.")
    else:
        print(f"No data rows in {SHEET}.")
    # Optionally reset other sheets? Not now.
    wb.save(BACKLOG_PATH)
    print("Backlog cleared. Header remains.")

if __name__ == "__main__":
    main()