#!/usr/bin/env python3
"""
Fix missing listing_url and mockup_url for today's rows.
"""
import openpyxl, datetime, subprocess, sys, os, time

BACKLOG_PATH = "/Users/duytan/Documents/Business/Printfamily/etsy-ideas.xlsx"
SHEET = "backlog"

def load_backlog():
    wb = openpyxl.load_workbook(BACKLOG_PATH)
    ws = wb[SHEET]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        rows.append(dict(zip(headers, row)))
    return rows

def update_row(row_id, updates):
    # Build command: etsy_backlog.py update <row_id> --field1 val1 ...
    cmd = [sys.executable, "etsy_backlog.py", "update", str(row_id)]
    for field, val in updates.items():
        if val is not None:
            cmd.extend([f"--{field.replace('_', '-')}", str(val)])
    try:
        result = subprocess.run(cmd, timeout=30, cwd=os.path.dirname(__file__), capture_output=True, text=True)
        if result.returncode != 0:
            print(f"  ❌ Update failed: {result.stderr[:100]}")
            return False
        return True
    except Exception as e:
        print(f"  ❌ Update error: {e}")
        return False

def main():
    rows = load_backlog()
    today = datetime.datetime.now().strftime("%Y-%m-%d")
    today_rows = [r for r in rows if r.get('spy_date') == today]
    print(f"Total rows today: {len(today_rows)}")
    
    updated = 0
    for row in today_rows:
        rid = row.get('id')
        lid = row.get('listing_id')
        if not lid:
            continue
        existing_url = row.get('listing_url')
        existing_mockup = row.get('mockup_url')
        updates = {}
        if not existing_url:
            updates['listing_url'] = f"https://www.etsy.com/listing/{lid}"
        if existing_mockup is None or existing_mockup == '':
            updates['mockup_url'] = ''
        if updates:
            print(f"Row {rid} (#{lid}) → adding {list(updates.keys())}")
            if update_row(rid, updates):
                updated += 1
                time.sleep(0.1)
            else:
                print(f"  Failed.")
    print(f"\n✅ Updated {updated} rows.")

if __name__ == "__main__":
    main()