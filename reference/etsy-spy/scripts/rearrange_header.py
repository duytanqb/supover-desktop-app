#!/usr/bin/env python3
"""
Rearrange backlog sheet header to match simplified FIELDNAMES.
"""
import openpyxl
from openpyxl.utils import get_column_letter

BACKLOG_PATH = "/Users/duytan/Documents/Business/Printfamily/etsy-ideas.xlsx"
SHEET = "backlog"

# New header order (must match FIELDNAMES in etsy_backlog.py)
NEW_HEADERS = [
    "id",
    "spy_date",
    "source_platform",
    "search_keyword",
    "source_query",
    "query_chain",
    "keyword_expansion",
    "save_reason",
    "niche",
    "listing_id",
    "listing_url",
    "title",
    "sold_24h",
    "views_24h",
    "hey_score",
    "days_old",
    "total_sold",
    "estimated_revenue",
    "tags",
    "trend_status",
    "trend_score",
    "design_angle",
    "workflow_status",
    "mockup_url",
    "updated_at",
]

def main():
    wb = openpyxl.load_workbook(BACKLOG_PATH)
    if SHEET not in wb.sheetnames:
        print(f"Sheet {SHEET} not found.")
        return
    
    ws = wb[SHEET]
    # Read existing data
    old_headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(old_headers, row)))
    
    # Create new sheet with temporary name
    new_ws = wb.create_sheet(title=f"{SHEET}_new")
    # Write new headers
    for col_idx, header in enumerate(NEW_HEADERS, start=1):
        new_ws.cell(row=1, column=col_idx, value=header)
    
    # Map old data to new headers
    for row_idx, old_row in enumerate(rows, start=2):
        for col_idx, header in enumerate(NEW_HEADERS, start=1):
            value = old_row.get(header)  # if header exists in old row
            new_ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Remove old sheet and rename new sheet
    del wb[SHEET]
    new_ws.title = SHEET
    wb.save(BACKLOG_PATH)
    print(f"Rearranged header. New columns: {len(NEW_HEADERS)}")
    print("Headers:", NEW_HEADERS)

if __name__ == "__main__":
    main()