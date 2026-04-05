#!/usr/bin/env python3
"""Shared Etsy spy backlog stored in Excel (.xlsx)."""

import argparse
import json
import re
import shutil
import zipfile
from datetime import datetime
from pathlib import Path
import xml.etree.ElementTree as ET

from openpyxl import Workbook, load_workbook

BACKLOG_PATH = Path("/Users/duytan/Documents/Business/Printfamily/etsy-ideas.xlsx")
SHEET_NAME = "backlog"
KEYWORD_SHEET_NAME = "keyword_log"
FRONTIER_SHEET_NAME = "frontier"
SEEN_SHEET_NAME = "seen"
CLUSTERS_SHEET_NAME = "clusters"
FIELDNAMES = [
    "id",
    "spy_date",
    "source_platform",
    "search_keyword",
    "source_query",
    "query_chain",
    "keyword_expansion",
    "save_reason",
    "niche",
    "listing_id",
    "listing_url",
    "title",
    "sold_24h",
    "views_24h",
    "hey_score",
    "days_old",
    "total_sold",
    "estimated_revenue",
    "tags",
    "trend_status",
    "trend_score",
    "design_angle",
    "workflow_status",
    "mockup_url",
    "updated_at",
]
KEYWORD_LOG_FIELDS = [
    "logged_at",
    "seed_keyword",
    "expanded_keyword",
    "source_type",
    "source_listing_id",
    "source_tag",
    "query_chain",
    "notes",
]
FRONTIER_FIELDS = [
    "queued_at",
    "status",
    "priority",
    "seed_keyword",
    "frontier_keyword",
    "source_type",
    "source_listing_id",
    "source_shop_name",
    "parent_query",
    "query_chain",
    "depth",
    "attempts",
    "last_seen_at",
    "notes",
]
SEEN_FIELDS = [
    "seen_at",
    "entity_type",
    "entity_id",
    "entity_value",
    "source_query",
    "source_listing_id",
    "source_shop_name",
    "cluster_key",
    "notes",
]
CLUSTER_FIELDS = [
    "cluster_key",
    "cluster_label",
    "seed_keyword",
    "representative_query",
    "status",
    "listing_count",
    "winner_count",
    "top_listing_id",
    "last_activity_at",
    "notes",
]
LEGACY_HEADER_ALIASES = {
    "expanded_keywords": "keyword_expansion",
    "expanded_keyword": "keyword_expansion",
    "expanded_query": "keyword_expansion",
    "reason_to_save": "save_reason",
    "save_note": "save_reason",
    "seed_query": "source_query",
    "keyword": "frontier_keyword",
    "query": "representative_query",
}
XML_NS = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
SHEET_FIELDS = {
    SHEET_NAME: FIELDNAMES,
    KEYWORD_SHEET_NAME: KEYWORD_LOG_FIELDS,
    FRONTIER_SHEET_NAME: FRONTIER_FIELDS,
    SEEN_SHEET_NAME: SEEN_FIELDS,
    CLUSTERS_SHEET_NAME: CLUSTER_FIELDS,
}


def remap_row(row: dict, allowed_fields=None) -> dict:
    mapped = {}
    allowed = set(allowed_fields) if allowed_fields else None
    for key, value in row.items():
        if key in (None, ""):
            continue
        canonical = LEGACY_HEADER_ALIASES.get(str(key), str(key))
        if allowed is None or canonical in allowed:
            mapped[canonical] = "" if value is None else str(value)
    return mapped


def salvage_rows_from_corrupt_workbook(path: Path):
    rows = []
    try:
        with zipfile.ZipFile(path) as zf:
            handle = zf.open("xl/worksheets/sheet1.xml")
            raw = handle.read(24000)
    except Exception:
        return rows

    text = raw.decode("utf-8", errors="ignore")
    start = text.find("<sheetData>")
    if start == -1:
        return rows

    marker = '<row r="5"'
    end = text.find("</row>", text.find(marker))
    if end == -1:
        return rows

    fragment = text[start:end + len("</row>")] + "</sheetData>"
    try:
        root = ET.fromstring(fragment)
    except ET.ParseError:
        return rows

    headers = []
    for row_el in root.findall("a:row", XML_NS):
        values = []
        for cell in row_el.findall("a:c", XML_NS):
            inline = cell.find("a:is", XML_NS)
            if inline is not None:
                value = "".join(t.text or "" for t in inline.iterfind(".//a:t", XML_NS))
            else:
                value = (cell.findtext("a:v", default="", namespaces=XML_NS) or "")
            values.append(value)
        if not headers:
            headers = values
            continue
        if any(v not in (None, "") for v in values):
            rows.append(remap_row(dict(zip(headers, values)), FIELDNAMES))
    return rows


def create_clean_workbook(path: Path, rows=None):
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(FIELDNAMES)
    for row in rows or []:
        normalized = {k: "" if v is None else str(v) for k, v in remap_row(row, FIELDNAMES).items()}
        ws.append([normalized.get(f, "") for f in FIELDNAMES])

    for sheet_name, fields in [
        (KEYWORD_SHEET_NAME, KEYWORD_LOG_FIELDS),
        (FRONTIER_SHEET_NAME, FRONTIER_FIELDS),
        (SEEN_SHEET_NAME, SEEN_FIELDS),
        (CLUSTERS_SHEET_NAME, CLUSTER_FIELDS),
    ]:
        ws2 = wb.create_sheet(sheet_name)
        ws2.append(fields)
    wb.save(path)
    wb.close()


def ensure_sheet(wb, sheet_name: str, fields):
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        ws.append(fields)
        return ws

    ws = wb[sheet_name]
    if ws.max_row == 0:
        ws.append(fields)
        return ws

    header_cells = list(ws[1]) if ws.max_row >= 1 else []
    current_headers = [cell.value for cell in header_cells[: len(fields)]]
    if current_headers != fields or ws.max_column != len(fields):
        existing = []
        headers = [cell.value for cell in ws[1]]
        header_index = {LEGACY_HEADER_ALIASES.get(h, h): i for i, h in enumerate(headers) if h}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(v not in (None, "") for v in row):
                continue
            item = {h: ("" if row[idx] is None else str(row[idx])) for h, idx in header_index.items() if idx < len(row)}
            existing.append(remap_row(item, fields))
        ws.delete_rows(1, ws.max_row)
        ws.append(fields)
        for item in existing:
            ws.append([item.get(f, "") for f in fields])
    return ws


def ensure_backlog(path: Path = BACKLOG_PATH):
    path.parent.mkdir(parents=True, exist_ok=True)
    if not path.exists():
        create_clean_workbook(path)
        return path

    try:
        wb = load_workbook(path)
    except Exception:
        salvaged = salvage_rows_from_corrupt_workbook(path)
        backup = path.with_suffix(path.suffix + ".corrupt")
        if not backup.exists():
            shutil.copy2(path, backup)
        create_clean_workbook(path, salvaged)
        return path

    for sheet_name, fields in SHEET_FIELDS.items():
        ensure_sheet(wb, sheet_name, fields)

    wb.save(path)
    wb.close()
    return path


def normalize(text: str) -> str:
    text = (text or "").strip().lower()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return re.sub(r"-+", "-", text).strip("-")


def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def build_concept_key(niche: str, title: str, design_angle: str) -> str:
    parts = [normalize(niche), normalize(title)[:60], normalize(design_angle)[:60]]
    return "|".join([p for p in parts if p])


def read_sheet_rows(sheet_name: str, path: Path = BACKLOG_PATH):
    ensure_backlog(path)
    wb = load_workbook(path)
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v not in (None, "") for v in row):
            continue
        rows.append({headers[i]: "" if row[i] is None else str(row[i]) for i in range(len(headers))})
    wb.close()
    return rows


def write_sheet_rows(sheet_name: str, fields, rows, path: Path = BACKLOG_PATH):
    ensure_backlog(path)
    wb = load_workbook(path)
    ws = wb[sheet_name]

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for row in rows:
        ws.append([row.get(f, "") for f in fields])

    wb.save(path)
    wb.close()


def read_rows(path: Path = BACKLOG_PATH):
    return read_sheet_rows(SHEET_NAME, path)


def write_rows(rows, path: Path = BACKLOG_PATH):
    write_sheet_rows(SHEET_NAME, FIELDNAMES, rows, path)


def next_id(rows) -> str:
    return str(max([int(r.get("id") or 0) for r in rows] + [0]) + 1)


def add_or_update(row: dict, path: Path = BACKLOG_PATH):
    row = remap_row(row, FIELDNAMES)
    rows = read_rows(path)
    listing_id = str(row.get("listing_id") or "").strip()
    concept_key = row.get("concept_key") or build_concept_key(row.get("niche", ""), row.get("title", ""), row.get("design_angle", ""))
    row["concept_key"] = concept_key
    row["updated_at"] = now_iso()
    row.setdefault("spy_date", datetime.now().strftime("%Y-%m-%d"))
    row.setdefault("source_platform", "etsy")
    row.setdefault("workflow_status", "spied")

    match = None
    for existing in rows:
        if listing_id and existing.get("listing_id") == listing_id:
            match = existing
            break

    if match:
        for key in FIELDNAMES:
            if key in row and row[key] not in (None, ""):
                match[key] = str(row[key])
        result = {"action": "updated", "row": match}
    else:
        new_row = {k: "" for k in FIELDNAMES}
        new_row.update({k: str(v) for k, v in row.items() if v is not None})
        new_row["id"] = next_id(rows)
        new_row["updated_at"] = now_iso()
        rows.append(new_row)
        result = {"action": "added", "row": new_row}

    write_rows(rows, path)
    return result


def upsert_sheet_row(sheet_name: str, fields, row: dict, match_keys, path: Path = BACKLOG_PATH):
    row = remap_row(row, fields)
    rows = read_sheet_rows(sheet_name, path)
    normalized = {k: "" if v is None else str(v) for k, v in row.items()}
    match = None
    for existing in rows:
        if all((existing.get(k) or "") == (normalized.get(k) or "") for k in match_keys):
            match = existing
            break

    if match:
        for key in fields:
            if key in normalized and normalized[key] not in (None, ""):
                match[key] = normalized[key]
        result = {"action": "updated", "row": match}
    else:
        new_row = {k: "" for k in fields}
        new_row.update(normalized)
        rows.append(new_row)
        result = {"action": "added", "row": new_row}

    write_sheet_rows(sheet_name, fields, rows, path)
    return result


def add_frontier_entry(row: dict, path: Path = BACKLOG_PATH):
    payload = {**row}
    payload.setdefault("queued_at", now_iso())
    payload.setdefault("status", "queued")
    payload.setdefault("priority", "normal")
    payload.setdefault("attempts", "0")
    payload.setdefault("last_seen_at", now_iso())
    return upsert_sheet_row(FRONTIER_SHEET_NAME, FRONTIER_FIELDS, payload, ["frontier_keyword", "query_chain"], path)


def add_seen_entry(row: dict, path: Path = BACKLOG_PATH):
    payload = {**row}
    payload.setdefault("seen_at", now_iso())
    return upsert_sheet_row(SEEN_SHEET_NAME, SEEN_FIELDS, payload, ["entity_type", "entity_id"], path)


def add_cluster_entry(row: dict, path: Path = BACKLOG_PATH):
    payload = {**row}
    payload.setdefault("last_activity_at", now_iso())
    payload.setdefault("status", "active")
    return upsert_sheet_row(CLUSTERS_SHEET_NAME, CLUSTER_FIELDS, payload, ["cluster_key"], path)


def filter_rows(rows, workflow_status=None, niche=None, limit=None):
    out = rows
    if workflow_status:
        out = [r for r in out if r.get("workflow_status") == workflow_status]
    if niche:
        out = [r for r in out if niche.lower() in (r.get("niche") or "").lower()]
    if limit:
        out = out[:limit]
    return out


def update_rows(ids, updates: dict, path: Path = BACKLOG_PATH):
    updates = remap_row(updates, FIELDNAMES)
    ids = {str(i) for i in ids}
    rows = read_rows(path)
    changed = []
    for row in rows:
        if row.get("id") in ids:
            for k, v in updates.items():
                if k in FIELDNAMES and v is not None:
                    row[k] = str(v)
            row["updated_at"] = now_iso()
            changed.append(row)
    write_rows(rows, path)
    return changed


def qualifies(row: dict) -> bool:
    sold_24h = float(row.get("sold_24h") or 0)
    views_24h = float(row.get("views_24h") or 0)
    hey = float(row.get("hey_score") or row.get("hey") or 0)
    days_old = float(row.get("days_old") or 999)
    return (
        sold_24h >= 2
        or views_24h >= 120
        or (views_24h >= 80 and hey >= 8)
        or (days_old <= 30 and hey >= 10 and views_24h >= 40)
        or (sold_24h >= 3 and days_old <= 90)
    )


def _numeric_listing_id(value):
    text = str(value or "").strip()
    return text if text.isdigit() else None


def known_listing_ids(path: Path = BACKLOG_PATH) -> set[str]:
    """Collect listing IDs already represented in workbook state.

    This is the cheap pre-dedupe set used before spending VK1ng / HEY API calls.
    It intentionally unions listing references across:
      - backlog rows (including excluded / rejected statuses)
      - seen sheet listing entities
      - frontier source listing references
      - cluster top_listing_id rollups
    """
    ids = set()

    for row in read_rows(path):
        listing_id = _numeric_listing_id(row.get("listing_id"))
        if listing_id:
            ids.add(listing_id)

    for row in read_sheet_rows(SEEN_SHEET_NAME, path):
        if (row.get("entity_type") or "").strip().lower() == "listing":
            listing_id = _numeric_listing_id(row.get("entity_id")) or _numeric_listing_id(row.get("source_listing_id"))
            if listing_id:
                ids.add(listing_id)

    for row in read_sheet_rows(FRONTIER_SHEET_NAME, path):
        listing_id = _numeric_listing_id(row.get("source_listing_id"))
        if listing_id:
            ids.add(listing_id)

    for row in read_sheet_rows(CLUSTERS_SHEET_NAME, path):
        listing_id = _numeric_listing_id(row.get("top_listing_id"))
        if listing_id:
            ids.add(listing_id)

    return ids


def filter_new_listing_ids(listing_ids, path: Path = BACKLOG_PATH):
    known = known_listing_ids(path)
    deduped = []
    seen_batch = set()
    skipped = []

    for listing_id in listing_ids:
        normalized = _numeric_listing_id(listing_id)
        if not normalized:
            continue
        if normalized in seen_batch or normalized in known:
            skipped.append(normalized)
            continue
        seen_batch.add(normalized)
        deduped.append(int(normalized))

    return deduped, skipped


def log_keyword_expansion(entry: dict, path: Path = BACKLOG_PATH):
    ensure_backlog(path)
    wb = load_workbook(path)
    ws = wb[KEYWORD_SHEET_NAME]
    payload = {k: "" if entry.get(k) is None else str(entry.get(k)) for k in KEYWORD_LOG_FIELDS}
    if not payload.get("logged_at"):
        payload["logged_at"] = now_iso()
    ws.append([payload.get(f, "") for f in KEYWORD_LOG_FIELDS])
    wb.save(path)
    wb.close()
    return payload


def print_rows(rows, as_json=False):
    if as_json:
        print(json.dumps(rows, indent=2, ensure_ascii=False))
        return
    if not rows:
        print("(no rows)")
        return
    for row in rows:
        title = row.get('title') or row.get('expanded_keyword') or row.get('frontier_keyword') or row.get('cluster_label') or row.get('entity_value') or ''
        status = row.get('workflow_status') or row.get('status') or row.get('entity_type') or ''
        print(f"{status} | {title[:70]}")
        print(f"   {json.dumps(row, ensure_ascii=False)}")


def add_state_parser(sub, name, fields, required_fields=None):
    parser = sub.add_parser(name)
    for field in fields:
        parser.add_argument(f"--{field.replace('_', '-')}", required=field in (required_fields or set()))
    return parser


def main():
    parser = argparse.ArgumentParser(description="Shared Etsy spy backlog (.xlsx)")
    parser.add_argument("--json", action="store_true")
    sub = parser.add_subparsers(dest="command")

    sub.add_parser("init")

    add = sub.add_parser("add")
    for field in [f for f in FIELDNAMES if f not in {"id", "updated_at"}]:
        add.add_argument(f"--{field.replace('_', '-')}")

    log_p = sub.add_parser("log-keyword")
    for field in KEYWORD_LOG_FIELDS:
        arg = f"--{field.replace('_', '-')}"
        if field == "expanded_keyword":
            log_p.add_argument(arg, required=True)
        else:
            log_p.add_argument(arg)

    add_state_parser(sub, "frontier-add", FRONTIER_FIELDS, {"frontier_keyword"})
    add_state_parser(sub, "seen-add", SEEN_FIELDS, {"entity_type", "entity_id"})
    add_state_parser(sub, "cluster-add", CLUSTER_FIELDS, {"cluster_key"})

    list_p = sub.add_parser("list")
    list_p.add_argument("--workflow-status")
    list_p.add_argument("--niche")
    list_p.add_argument("--limit", type=int)

    state_list = sub.add_parser("state-list")
    state_list.add_argument("sheet", choices=[FRONTIER_SHEET_NAME, SEEN_SHEET_NAME, CLUSTERS_SHEET_NAME, KEYWORD_SHEET_NAME, SHEET_NAME])
    state_list.add_argument("--limit", type=int)

    next_p = sub.add_parser("next")
    next_p.add_argument("--workflow-status", default="approved")
    next_p.add_argument("--limit", type=int, default=1)

    upd = sub.add_parser("update")
    upd.add_argument("ids", nargs="+")
    for field in [f for f in FIELDNAMES if f not in {"id"}]:
        upd.add_argument(f"--{field.replace('_', '-')}")

    args = parser.parse_args()

    if args.command == "init":
        path = ensure_backlog()
        print(path)
        return

    if args.command == "add":
        payload = {}
        for field in [f for f in FIELDNAMES if f not in {"id", "updated_at"}]:
            val = getattr(args, field)
            if val is not None:
                payload[field] = val
        if not payload.get("concept_key"):
            payload["concept_key"] = build_concept_key(payload.get("niche", ""), payload.get("title", ""), payload.get("design_angle", ""))
        result = add_or_update(payload)
        print_rows([result["row"]], as_json=args.json)
        return

    if args.command == "log-keyword":
        payload = {field: getattr(args, field) for field in KEYWORD_LOG_FIELDS if getattr(args, field) is not None}
        result = log_keyword_expansion(payload)
        print_rows([result], as_json=args.json)
        return

    if args.command == "frontier-add":
        payload = {field: getattr(args, field) for field in FRONTIER_FIELDS if getattr(args, field) is not None}
        result = add_frontier_entry(payload)
        print_rows([result["row"]], as_json=args.json)
        return

    if args.command == "seen-add":
        payload = {field: getattr(args, field) for field in SEEN_FIELDS if getattr(args, field) is not None}
        result = add_seen_entry(payload)
        print_rows([result["row"]], as_json=args.json)
        return

    if args.command == "cluster-add":
        payload = {field: getattr(args, field) for field in CLUSTER_FIELDS if getattr(args, field) is not None}
        result = add_cluster_entry(payload)
        print_rows([result["row"]], as_json=args.json)
        return

    if args.command == "list":
        rows = filter_rows(read_rows(), args.workflow_status, args.niche, args.limit)
        print_rows(rows, as_json=args.json)
        return

    if args.command == "state-list":
        rows = read_sheet_rows(args.sheet)
        if args.limit:
            rows = rows[:args.limit]
        print_rows(rows, as_json=args.json)
        return

    if args.command == "next":
        rows = filter_rows(read_rows(), args.workflow_status, None, args.limit)
        print_rows(rows, as_json=args.json)
        return

    if args.command == "update":
        updates = {}
        for field in [f for f in FIELDNAMES if f != "id"]:
            val = getattr(args, field)
            if val is not None:
                updates[field] = val
        rows = update_rows(args.ids, updates)
        print_rows(rows, as_json=args.json)
        return

    parser.print_help()


if __name__ == "__main__":
    main()
