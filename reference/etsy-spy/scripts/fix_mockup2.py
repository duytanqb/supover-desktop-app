#!/usr/bin/env python3
"""
Set mockup_url = '' for rows 30 onward.
"""
import openpyxl, subprocess, sys, os, time

BACKLOG_PATH = "/Users/duytan/Documents/Business/Printfamily/etsy-ideas.xlsx"
SHEET = "backlog"

def load_backlog():
    wb = openpyxl.load_workbook(BACKLOG_PATH)
    ws = wb[SHEET]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        rows.append(dict(zip(headers, row)))
    return rows

def update_row(row_id, updates):
    cmd = [sys.executable, "etsy_backlog.py", "update", str(row_id)]
    for field, val in updates.items():
        if val is not None:
            cmd.extend([f"--{field.replace('_', '-')}", str(val)])
    try:
        result = subprocess.run(cmd, timeout=30, cwd=os.path.dirname(__file__), capture_output=True, text=True)
        if result.returncode != 0:
            print(f"  ❌ Update failed: {result.stderr[:100]}")
            return False
        return True
    except Exception as e:
        print(f"  ❌ Update error: {e}")
        return False

def main():
    rows = load_backlog()
    print(f"Total rows: {len(rows)}")
    
    updated = 0
    for row in rows:
        rid = int(row.get('id'))
        if rid < 30:
            continue
        lid = row.get('listing_id')
        existing_mockup = row.get('mockup_url')
        if existing_mockup is None or existing_mockup == '' or existing_mockup == 'None':
            print(f"Row {rid} (#{lid}) → setting mockup_url=''")
            if update_row(rid, {'mockup_url': ''}):
                updated += 1
                time.sleep(0.1)
            else:
                print(f"  Failed.")
    print(f"\n✅ Updated {updated} rows.")

if __name__ == "__main__":
    main()