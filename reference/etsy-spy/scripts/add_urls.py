#!/usr/bin/env python3
"""
Add listing_url and mockup_url to today's backlog rows.
"""
import openpyxl, datetime, subprocess, sys, os, time

BACKLOG_PATH = "/Users/duytan/Documents/Business/Printfamily/etsy-ideas.xlsx"
SHEET = "backlog"

def load_backlog():
    wb = openpyxl.load_workbook(BACKLOG_PATH)
    ws = wb[SHEET]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        rows.append(dict(zip(headers, row)))
    return rows

def update_row(row_id, updates):
    cmd = [sys.executable, "etsy_backlog.py", "update", str(row_id)]
    for field, val in updates.items():
        if val is not None and val != "":
            cmd.extend([f"--{field.replace('_', '-')}", str(val)])
    try:
        subprocess.run(cmd, timeout=30, cwd=os.path.dirname(__file__), capture_output=True)
        return True
    except Exception as e:
        print(f"  Update error: {e}")
        return False

def main():
    rows = load_backlog()
    today = datetime.datetime.now().strftime("%Y-%m-%d")
    today_rows = [r for r in rows if r.get('spy_date') == today]
    print(f"Total rows today: {len(today_rows)}")
    
    updated = 0
    for row in today_rows:
        rid = row.get('id')
        lid = row.get('listing_id')
        existing_url = row.get('listing_url')
        existing_mockup = row.get('mockup_url')
        updates = {}
        if not existing_url and lid:
            updates['listing_url'] = f"https://www.etsy.com/listing/{lid}"
        if not existing_mockup:
            updates['mockup_url'] = ""  # empty string
        if updates:
            print(f"Row {rid} (#{lid}) → adding {list(updates.keys())}")
            if update_row(rid, updates):
                updated += 1
                time.sleep(0.2)
            else:
                print(f"  Failed.")
    print(f"\n✅ Updated {updated} rows.")

if __name__ == "__main__":
    main()